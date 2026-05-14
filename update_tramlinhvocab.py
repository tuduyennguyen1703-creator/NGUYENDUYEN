import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_tramlinh_template(excel_file):
    path = Path(excel_file)
    if path.exists():
        print(f"File '{excel_file}' already exists. Skipping template creation.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tramlinh Vocab'

    headers = ['STT', 'English', 'Pos', 'IPA', 'Vietnamese', 'Example']
    ws.append(headers)
    sample_row = [1, 'Mental clarity', 'n. phr.', '/ˌmen.təl ˈklær.ə.ti/', 'Sự minh mẫn, sáng suốt về tinh thần', 'Your mental clarity improves when you meditate daily.']
    ws.append(sample_row)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4B6584')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = [7, 32, 18, 26, 34, 46]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 24

    wb.security.set_workbook_password('2009')
    wb.save(excel_file)
    print(f"Created template file '{excel_file}' with password 2009.")


def update_tramlinh_js(excel_file, js_file):
    print(f"Reading data from '{excel_file}'...")
    df = pd.read_excel(excel_file, header=None).fillna("")

    first_cell = str(df.iloc[0, 0]).strip().lower()
    if first_cell in ['num', 'id', 'stt']:
        df = df[1:].reset_index(drop=True)

    col_names = ['num', 'en', 'pos', 'ipa', 'vi', 'ex'] # Default column names for tramlinhvocab.js
    
    # Apply robust column renaming and adding/dropping logic
    temp_df = pd.DataFrame(columns=col_names)
    for i, col_name in enumerate(col_names):
        if i < df.shape[1]:
            temp_df[col_name] = df.iloc[:, i]
        else:
            temp_df[col_name] = "" # Add missing columns as empty
    df = temp_df.copy() # Use the new DataFrame with correct columns and names

    # Ensure 'num' column is integer
    if 'num' in df.columns:
        df['num'] = pd.to_numeric(df['num'], errors='coerce').fillna(0).astype(int)


    vocab_list = df.to_dict(orient='records')
    with open(js_file, 'w', encoding='utf-8') as f:
        f.write('const vocabularyList = ')
        json.dump(vocab_list, f, ensure_ascii=False, indent=4)
        f.write(';\n')

    print(f"Updated {len(vocab_list)} vocab entries into '{js_file}'.")


if __name__ == '__main__':
    excel_path = 'tu_vung_tramlinh.xlsx'
    js_path = 'tramlinhvocab.js'
    create_tramlinh_template(excel_path)
    update_tramlinh_js(excel_path, js_path)
